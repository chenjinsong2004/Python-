from django.shortcuts import render, redirect
from app01 import models
from app01.utils.pagination import Pagination
from django.http import HttpResponse
from app01.utils.form import UserModelForm, PrettyEditModelForm, PrettyModelForm, AdminModelForm, AdminEditModelForm, AdminResetModelForm
from openpyxl import load_workbook
# Create your views here.
def depart_list(request):

    queryset = models.Department.objects.all()

    return render(request, 'depart_list.html', {"depart_list": queryset})

def depart_add(request):
    if request.method == "GET":
        return render(request, 'depart_add.html')
    # 获取用户输入信息
    title = request.POST.get("title")
    # 保存到数据库
    models.Department.objects.create(title=title)
    # 重定向会部门列表
    return redirect("/depart/list/")

def depart_delete(request):
    # http://127.0.0.1:8000/depart/delete/?nid=1
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")

def depart_edit(request, nid):
    # 根据id，获取他的数据[obj, ]
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        print(row_object.id, row_object.title)
        return render(request, 'depart_edit.html', {"row_object":row_object})
    
    # 获取用户提交的标题
    title = request.POST.get("title")
    # 根据id找数据库的数据进行更新
    models.Department.objects.filter(id=nid).update(title=title)
    # 重定向
    return redirect("/depart/list/")

def depart_multi(request):
    # 批量上传
    # 获取用户上传的文件对象
    file_object = request.FILES.get('exc')

    # 直接打开Excel并读取内容
    
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    cell = sheet.cell(1,1)
    # print(cell.value)

    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        print(text)
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return HttpResponse('上传文件')

def user_list(request):
    queryset = models.UserInfo.objects.all()
    page_object = Pagination(request, queryset, page_size=1)
    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()
    }
    if request.method == "GET":
        return render(request, 'user_list.html',context)
def user_add(request):
    if request.method == "GET":
        context = {
            'gender_choice': models.UserInfo.gender_choice,
            'department': models.Department.objects.all()
        }
        return render(request, 'user_add.html', context)
    
    # 获取用户提交数据
    user = request.POST.get('name')
    age = request.POST.get('age')
    account = request.POST.get('account')
    create_time = request.POST.get('create_time')
    depart = request.POST.get('depart')
    gender = request.POST.get('gender')

    # 添加到数据库中
    models.UserInfo.objects.create(name=user,age=age,account=account,create_time=create_time,depart_id=depart,gender=gender)
    return redirect('/user/list')

# 用户 ModelForm 添加视图
def user_model_form_add(request):
    if request.method == "GET":
        form = UserModelForm()
        return render(request, 'user_model_form_add.html', {"form": form})
    # 用户POST提交数据，数据校验
    form = UserModelForm(data=request.POST)
    # 校验
    if form.is_valid():
        form.save()
        return redirect('/user/list')
    else:
        return render(request, 'user_model_form_add.html', {"form": form})
    
def user_edit(request, nid):

    row_object = models.UserInfo.objects.filter(id=nid).first()

    if request.method == "GET":
        form = UserModelForm(instance=row_object)
        return render(request, 'user_edit.html', {'form':form})
    
    form = UserModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/user/list/')
    return render(request, 'user_edit.html', {'form':form})

def user_delete(request, nid):
    models.UserInfo.objects.filter(id=nid).delete()
    return redirect('/user/list')

# 靓号
def pretty_list(request):
    # print(request.GET)
    info  = request.session.get("info")
    if not info:
        return redirect("/login/")

    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["mobile__contains"] = search_data
    queryset = models.MobileInfo.objects.filter(**data_dict).order_by("-level")
    # 根据用户想要的页面计算出数据
    page_object = Pagination(request, queryset)

    context = {
        "queryset":page_object.page_queryset,  # 分完页的数据
        "search_data":search_data,
        "page_string":page_object.html(), # 页码
    }

    return render(
        request,
        'pretty_list.html',
        context,
    )

def pretty_add(request):
    if request.method == "GET":
        form = PrettyModelForm()
        return render(request, 'pretty_add.html', {"form":form})
    # 用户POST提交数据，数据校验
    form = PrettyModelForm(data=request.POST)
    # 校验
    if form.is_valid():
        form.save()
        return redirect('/pretty/list/')
    else:
        return render(request, 'pretty_add.html', {"form": form})

def pretty_edit(request, nid):
    row_object = models.MobileInfo.objects.filter(id=nid).first()

    if request.method == "GET":
        form = PrettyEditModelForm(instance=row_object)
        return render(request, 'pretty_edit.html', {'form':form})
    
    form = PrettyEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/pretty/list/')
    return render(request, 'pretty_edit.html', {'form':form})

def pretty_delete(request, nid):

    models.MobileInfo.objects.filter(id=nid).delete()
    return redirect('/pretty/list')

# 管理员
def admin_list(request):
    info  = request.session.get("info")
    if not info:
        return redirect("/login/")

    # for i in 100:
    #     models.Admin.objects.create(
    #         username="张三",    # 用户名
    #         password="123456"   # 密码
    #     )
    # 搜索
    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["username__contains"] = search_data
    # 根据条件去数据库获取
    queryset = models.Admin.objects.filter(**data_dict)

    # 分页
    page_object = Pagination(request, queryset)
    context = {
        'queryset': queryset,
        'page_object': page_object,
        'search_data':search_data
    }
    if request.method == "GET":
        return render(request, 'admin_list.html', context)
    
def admin_add(request):
    if request.method == "GET":
        form = AdminModelForm()
        return render(request, 'change.html', {"form":form, "title":'添加管理员'})
    # 用户POST提交数据，数据校验
    form = AdminModelForm(data=request.POST)
    # 校验
    if form.is_valid():
        form.save()
        return redirect('/myadmin/list/')
    else:
        return render(request, 'change.html', {"form": form})
    
def admin_edit(request, nid):
    title = '修改管理员'
    row_object = models.Admin.objects.filter(id=nid).first()
    if not row_object:
        return redirect('/myadmin/list/')
    
    if request.method == "GET":
        form = AdminEditModelForm(instance=row_object)
        return render(request, 'change.html', {"form":form, "title":title, "nid":nid})
    
    # 用户POST提交数据，数据校验
    form = AdminEditModelForm(data=request.POST, instance=row_object)
    # 校验
    if form.is_valid():
        form.save()
        return redirect('/myadmin/list/')
    else:
        return render(request, 'change.html', {"form": form, "title":title})

def delete_edit(request, nid):
    models.Admin.objects.filter(id=nid).delete()
    return redirect('/myadmin/list/')

def admin_reset(request, nid):
    row_object = models.Admin.objects.filter(id=nid).first()
    if not row_object:
        return redirect('/myadmin/list/')
    
    title = "重置密码-{}".format(row_object.username)
    if request.method == "GET":
        form = AdminResetModelForm()
        return render(request, 'change.html', {"form":form, "title": title})
    
    form = AdminResetModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/myadmin/list/')
    return render(request, 'change.html', {"form":form, "title": title})